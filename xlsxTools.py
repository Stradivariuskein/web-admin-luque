from openpyxl import load_workbook
import os
import re
from datetime import datetime
from configs import *


ABC = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
# si el porcentaje es positivo aumenta sino hace un decuento
# retorna el numero con el porsentaje aplicado
def percent_apli(num, percent):
    aux = abs(percent / 100)
    apli_p = num * aux

    if percent > 0:
        result = num + apli_p
    elif percent < 0:
        result = num - apli_p
    else:
        result = 0

    return round(result,2)



#verifica si es un numero
def es_numero(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def buscarPrecio(cod, lista_num):
    #BUSCA POR CODIGO EL PRECIO DEL ARTICULO en siaac
    
    long_precio = 11
    if lista_num == 1:
        inicioPrecio = 20
    if lista_num == 5:
        inicioPrecio = 66

    finprecio = inicioPrecio + long_precio
    patron = '^' + cod.upper().strip() + ' '
    file = open("DB//articDB.txt")
    #bandera = 0
    i=0
    
    
    for linea in file:
       
        result = re.findall(patron, linea)
        #i+=1
        
        #if i == 4909 and cod == 'TUA-104':
        #    print(linea)
       
        if result:
            precio = linea[inicioPrecio:finprecio]
           
            return precio.strip(" ")

    print(f"\n\n*********************************************\n ERROR: codigo {cod} no encontrado reviselo\n*********************************************\n")    
    return -1


def actualizarPrecio (wb,row,cod,lista_num):
    #ACTUALIZA EL PRECIO EN EL EXCEL
    
    sheet = wb['Hoja1']
    cell = 'D' + str(row)
    precio = buscarPrecio(cod, lista_num)
    if precio != -1:
        sheet[cell] = float(precio)


def actualizarLista(bExcel, lista_num):
    #recorre una lista de precios, obtiene los codigo, los busca en articDB.txt y actualiza el precio
    try:
        sh = bExcel['Hoja1']
    except:
        print ("Error no exixte la Hoja1 en el archivo excel!!!!")
   
    cell=""
    i=0
    columna = 1

    now = datetime.now()    
    sh['A1'] = now

    maxFila = sh.max_row

    for i in range(1,maxFila):
        cell=str(sh.cell(row=i,column=columna).value).upper()
    
        celda = str(cell).upper()
        
        
        if celda == "COD" or celda == "COD.":
            i += 1
            cell=str(sh.cell(row=i,column=columna).value).upper()
            result = re.findall(f"[A-Z]-", cell)

            while result:
                precioActual = str(sh.cell(row=i,column=columna+3).value)
                if es_numero(precioActual):
                    actualizarPrecio(bExcel,i,cell,lista_num)
                    
                i+=1
                cell=sh.cell(row=i,column=columna).value
                if cell:                    
                    result = re.findall(f"[A-Z]-", cell.upper())
                else:
                    result = 0


# dado un archivo xlsx devuelve un dic con todos los codgos descripcion, precios, y cleda
def get_artcis_from_xlsx(rute_xlsx):
    
    try:
        wb = load_workbook(rute_xlsx)
    except FileNotFoundError:
        return None
    sheet = wb['Hoja1']

    maxRow = sheet.max_row
    
    code_pattern = "^[A-Z]+-"

    list_codes = {}
    col = 1
    row = 1
    while row < maxRow:
        cell=str(sheet.cell(row=row,column=col).value).upper()
        cell_title = cell
        if cell.strip() == 'T-246':
            print(cell)
        if cell_title == "COD" or cell_title == "COD.":
            row += 1
            cell=str(sheet.cell(row=row,column=col).value).upper()
            
            # valido q sea un codigo. el codigo se compone de letras mayusculas seguidas por un guion (ej: TIR-250)
            result = re.findall(code_pattern, cell)
            
            while result:

                
                price = sheet.cell(row=row,column=col+3).value
                if isinstance(price, str):
                    try:
                        price = float(price.strip())
                    except ValueError:
                        price = None
                list_codes[cell] = {
                    'description': str(sheet.cell(row=row,column=col+1).value).upper(),
                    'price': price,
                    'row': row,
                    'col': col
                }
                

                row += 1
                cell=str(sheet.cell(row=row,column=col).value).strip().upper()
                if cell:                    
                    result = re.findall(code_pattern, cell.upper())
                else:
                    result = None
        if cell == 'COD' or cell == 'COD.':
            pass
        else:
            row += 1

    return list_codes

# actualiza el archio xlsx con los precios el sistema (siaac)
def update_xlsx(xlsx_name, xlsx_data):

    for rute_xlsx in xlsx_data["rutes"]:

        try:
            wb = load_workbook(rute_xlsx)
        except FileNotFoundError:
            return None
        
        to_return = []
        sheet = wb['Hoja1']

        
        for code, data in xlsx_data.items():
            
            
            if code != 'rutes':
                try:
                    row = data['row']
                    col = data['col'] - 1
                    code = sheet[row][col].value 
                    code = code.strip()

                    if code:
                        cell = f"{ABC[col+4]}{row}"
                        percent = int(data['price_percent'])
                        if percent != 0:
                            xlsx_data[code]['percent'] = percent
                            try:  
                                cell_value = sheet[row][col+3].value
                                cell_value = float(cell_value)
                                result_mi = rute_xlsx.find('mi')
                                result_ma = rute_xlsx.find('ma')
                                if result_mi > -1:
                                    print(percent_apli(cell_value, percent))
                                    xlsx_data[code]['price_manual_min'] = percent_apli(cell_value, percent)
                                    sheet[cell] = xlsx_data[code]['price_manual_min']
                                elif result_ma > -1:
                                    xlsx_data[code]['price_manual_may'] = percent_apli(cell_value, percent)        
                                    sheet[cell] = xlsx_data[code]['price_manual_may']
                                
                            except ValueError:
                                print(f"error: en {code} el precio {cell_value} no es un numero")
                        elif data['price_manual_may'] != None and data['price_manual_min'] != None:
                            try:
                                is_ma = rute_xlsx.find("/ma/")
                                is_mi = rute_xlsx.find("/mi/")
                                if is_ma > -1:
                                    if isinstance(data['price_manual_may'], float):
                                        xlsx_data[code]['price_manual_may'] = data['price_manual_may']
                                    else:
                                        xlsx_data[code]['price_manual_may'] = float(data['price_manual_may'].strip())

                                    sheet[cell] = xlsx_data[code]['price_manual_may']
                                elif is_mi > -1:
                                    if isinstance(data['price_manual_min'], float):
                                        xlsx_data[code]['price_manual_min'] = data['price_manual_min']
                                    else:
                                        xlsx_data[code]['price_manual_min'] = float(data['price_manual_min'].strip())
                                    sheet[cell] = xlsx_data[code]['price_manual_min']
                            except ValueError:
                                
                                sheet[cell] = '********'
                            
                        else:
                            print(f"{code}: el precio no cambio")
                except Exception as e:
                    print(f"Error: {e}")
                to_return.append((code, xlsx_data[code]))

            
        sheet["A1"] = datetime.now().date()
        wb.save(rute_xlsx)

    return (xlsx_name, to_return)

def list_xlsx_to_folder(path):
    archivos_xlsx = []

    for root, dirs, files in os.walk(path):
        for file in files:
            if file.endswith(".xlsx"):
                archivos_xlsx.append((file, os.path.basename(root)))

    return archivos_xlsx


if __name__ == '__main__':

    

    results = list_xlsx_to_folder(RUTE_XLSX_AGRUPS['ma'])
    

    print(results)



    
    
