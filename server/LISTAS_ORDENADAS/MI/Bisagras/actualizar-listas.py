from openpyxl import load_workbook


import os
import re
import shutil

from progress.bar import Bar

from datetime import datetime

#verifica si es un numero
def es_numero(num):
    try:
        float(num)
        return True
    except ValueError:
        return False



def leerArtic():
    #PASA EL ARCHIVO DE LOS ARTICULOS DE SISTEMA A UN ARCHIVO DE TEXTO FACIL DE LEER
    shutil.copy(R"Y:/SIAAC3/ARTIC.DBF",R"./DB/ARTIC.DBF")
    file = open(R"DB/ARTIC.DBF", errors="ignore")
    articdb = open("DB//articDB.txt", "w")
    for i in range(0,6):
    #se descarta las pimeras lineas
        linea = file.readline()


    linea = file.readline(2)
    linea = file.readline(200)
    while linea != "":
        
        result = re.findall("[A-Z]$", linea)
        if result:
            lineaAux = linea[199]
            linea = linea[:-132] + " " + linea[68:-1] + '\n'
            articdb.write(linea[:-189].lstrip('\x00').lstrip() + linea[89:])
            linea = lineaAux + file.readline(199)
        else:
            articdb.write(linea[:-189].lstrip('\x00').lstrip() + linea[89:] + '\n')
            linea = file.readline(200)
    file.close()
    articdb.close()

def buscarPrecio(cod, lista_num):
    #BUSCA POR CODIGO EL PRECIO DEL ARTICULO 
    
    long_precio = 11
    if lista_num == 1:
        inicioPrecio = 20
    if lista_num == 5:
        inicioPrecio = 66

    finprecio = inicioPrecio + long_precio
    patron = '^' + cod.upper().strip() + ' '
    file = open("DB//articDB.txt")
    #bandera = 0
    i=0
    
    
    for linea in file:
       
        result = re.findall(patron, linea)
        #i+=1
        
        #if i == 4909 and cod == 'TUA-104':
        #    print(linea)
       
        if result:
            precio = linea[inicioPrecio:finprecio]
           
            return precio.strip(" ")

    print(f"\n\n*********************************************\n ERROR: codigo {cod} no encontrado reviselo\n*********************************************\n")    
    return -1


def actualizarLista(bExcel, lista_num):
    #recorre una lista de precios, obtiene los codigo, los busca en articDB.txt y actualiza el precio
    try:
        sh = bExcel['Hoja1']
    except:
        print ("Error no exixte la Hoja1 en el archivo excel!!!!")
   
    cell=""
    i=0
    columna = 1

    now = datetime.now()    
    sh['A1'] = now

    maxFila = sh.max_row

    for i in range(1,maxFila):
        cell=str(sh.cell(row=i,column=columna).value).upper()
    
        celda = str(cell).upper()
        
        
        if celda == "COD" or celda == "COD.":
            i += 1
            cell=str(sh.cell(row=i,column=columna).value).upper()
            result = re.findall(f"[A-Z]-", cell)

            while result:
                precioActual = str(sh.cell(row=i,column=columna+3).value)
                if es_numero(precioActual):
                    actualizarPrecio(bExcel,i,cell,lista_num)
                    
                i+=1
                cell=sh.cell(row=i,column=columna).value
                if cell:                    
                    result = re.findall(f"[A-Z]-", cell.upper())
                else:
                    result = 0



def actualizarPrecio (wb,row,cod,lista_num):
    #ACTUALIZA EL PRECIO EN EL EXCEL
    
    sheet = wb['Hoja1']
    cell = 'D' + str(row)
    precio = buscarPrecio(cod, lista_num)
    if precio != -1:
        sheet[cell] = float(precio)
    

def getListas():
    #CREA UNA LISTA CON TODOS LOS ARCHVOS .XLSX EN EL MISMO DIRECTORIO DE EJECUCON
    archivos = os.listdir('./')
    lista = []
    
    for line in archivos:
        result = re.findall("\S.xlsx", line)
        if result:
            lista.append(line)
            
    return lista





if __name__ == '__main__':

    leerArtic()
    
    listas =getListas()

    num_listas = len(listas)

    bar1 = Bar("Actualizando listas:", max=num_listas)
    lista_ma_mi = ""
    lista_num = 0
    msj = ""

    while lista_ma_mi == "":
        os.system("cls")

        lista_ma_mi = input(f"{msj}Ingrese 'MA' para mayorista o 'MI' para minorista: ")

        if lista_ma_mi.upper() == 'MI':
            lista_num = 1
        elif lista_ma_mi.upper() == 'MA':
            lista_num = 5
        else:
            lista_ma_mi = ""
            msj = "Opcion no valida\n"
    
    for arch in listas:
        try:
            book = load_workbook(filename= arch)
        except:
            print(F"\n\nERROR NO SE PUEDE ABRIR LA LISTA{arch}\nINTENTE CERRAR LAS LISTA E INTETELO NUEVAMENTE")
            
        
        actualizarLista(book, lista_num)

        
        try:
            book.save(f"./{arch}")
        except:
            print(f"\n\nError no se puedo guardar el archivo {arch}. intetnte cerrar todos los archivos excel e intentelo nuevamente")
        
        bar1.next()
    bar1.finish()

    input("\n\n--PRECIONE ENTER PARA SALIR--")

    

    
