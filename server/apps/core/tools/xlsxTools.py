from openpyxl import load_workbook
import os
import shutil
import re
from datetime import datetime
from configs import *
from apps.core.models import ModelArtic, ModelToUpdateList, ModelFileDrive

from django.db import transaction



ABC = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
# si el porcentaje es positivo aumenta sino hace un decuento
# retorna el numero con el porsentaje aplicado
def percent_apli(num, percent):
    aux = abs(percent / 100)
    apli_p = num * aux

    if percent > 0:
        result = num + apli_p
    elif percent < 0:
        result = num - apli_p
    else:
        result = 0

    return round(result,2)


def buscarPrecio(cod, lista_num):
    #BUSCA POR CODIGO EL PRECIO DEL ARTICULO en siaac
    
    long_precio = 11
    if lista_num == 1:
        inicioPrecio = 20
    if lista_num == 5:
        inicioPrecio = 66

    finprecio = inicioPrecio + long_precio
    patron = '^' + cod.upper().strip() + ' '
    file = open("DB//articDB.txt")
    #bandera = 0
    i=0
    
    
    for linea in file:
       
        result = re.findall(patron, linea)

       
        if result:
            precio = linea[inicioPrecio:finprecio]
           
            return precio.strip(" ")

    print(f"\n\n*********************************************\n ERROR: codigo {cod} no encontrado reviselo\n*********************************************\n")    
    return -1


# dado una ruta de un archivo xlsx devuelve un dic con todos los codgos descripcion, precios, y cleda
def get_artcis_from_xlsx(rute_xlsx):
    
    try:
        wb = load_workbook(rute_xlsx)
    except FileNotFoundError:
        print(f'{rute_xlsx} not found')
        return None
    sheet = wb['Hoja1']

    maxRow = sheet.max_row
    
    code_pattern = "^[A-Z]+-"

    list_codes = {}
    col = 1
    row = 1
    while row < maxRow:
        cell=str(sheet.cell(row=row,column=col).value).upper()
        cell_title = cell
        if cell_title == "COD" or cell_title == "COD.":
            row += 1
            cell=str(sheet.cell(row=row,column=col).value).upper()
            
            # valido q sea un codigo. el codigo se compone de letras mayusculas seguidas por un guion (ej: TIR-250)
            result = re.findall(code_pattern, cell)
            
            while result:

                
                price = sheet.cell(row=row,column=col+3).value
                if isinstance(price, str):
                    try:
                        price = float(price.strip())
                    except ValueError:
                        price = None
                list_codes[cell] = {
                    'description': str(sheet.cell(row=row,column=col+1).value).upper(),
                    'price': price,
                    'row': row,
                    'col': col
                }
                

                row += 1
                cell=str(sheet.cell(row=row,column=col).value).strip().upper()
                if cell:                    
                    result = re.findall(code_pattern, cell.upper())
                else:
                    result = None
        if cell == 'COD' or cell == 'COD.':
            pass
        else:
            row += 1

    return list_codes

def copy_file(origin, folder):
    # Obtener el nombre del archivo desde la ruta de origen
    file_name = os.path.basename(origin)
    
    # Crear la ruta completa de destino para el archivo
    dest_file = os.path.join(folder, file_name)

    # Verificar si ya existe un archivo con el mismo nombre en la carpeta destino
    if os.path.exists(dest_file):
        try:
            # Eliminar el archivo existente en la carpeta destino
            os.remove(dest_file)
            
        except FileNotFoundError:
            print(f"Error: No se pudo encontrar el archivo {dest_file}.")
        except PermissionError:
            print(f"Error: Permiso denegado al eliminar el archivo en {dest_file}.")
        except Exception as e:
            print(f"Error inesperado al eliminar el archivo: {e}")

    try:
        # Copiar el archivo desde el origen hasta la carpeta destino
        shutil.copy(os.path.join(origin), folder)
        #print(f"Archivo copiado de {origin} a {folder} exitosamente.")
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {origin}.")
    except PermissionError:
        print(f"Error: Permiso denegado al copiar el archivo en {folder}.")
    except Exception as e:
        print(f"Error inesperado al copiar el archivo: {e}")

# actualiza el archio xlsx con los precios el sistema (siaac)
def update_xlsx(xlsx_name, xlsx_data):

    for rute_xlsx in xlsx_data["rutes"]:

        try:
            wb = load_workbook(rute_xlsx)
        except FileNotFoundError:
            # hay q arreglar q cuando 1 ruta no exista crashea toda la app
            print(f"none: {rute_xlsx}")
            continue
        
        to_return = []
        sheet = wb['Hoja1']

        
        for code, data in xlsx_data.items():
            
            if code != 'rutes':
                try:
                    row = data['row']
                    col = data['col'] - 1
                    code = sheet[row][col].value 
                    code = code.strip()
                    
                    if code:
                        cell = f"{ABC[col+4]}{row}"
                        percent = int(data['price_percent'])
                        xlsx_data[code]['description'] = sheet[row][col-4].value
                        if percent != 0:
                            xlsx_data[code]['percent'] = percent
                            try:  
                                cell_value = sheet[row][col+3].value
                                cell_value = float(cell_value)
                                result_mi = rute_xlsx.upper().find('MI')
                                result_ma = rute_xlsx.upper().find('MA')
                                if result_mi > -1:
                                    xlsx_data[code]['price_manual_min'] = percent_apli(cell_value, percent)
                                    sheet[cell] = xlsx_data[code]['price_manual_min']
                                elif result_ma > -1:
                                    xlsx_data[code]['price_manual_may'] = percent_apli(cell_value, percent)        
                                    sheet[cell] = xlsx_data[code]['price_manual_may']
                                
                            except ValueError:
                                print(f"error: en {code} el precio {cell_value} no es un numero")
                        elif data['price_manual_may'] != None and data['price_manual_min'] != None:
                            try:
                                is_ma = rute_xlsx.upper().find("/MA/")
                                is_mi = rute_xlsx.upper().find("/MI/")

                                if is_ma > -1:
                                    if isinstance(data['price_manual_may'], float):
                                        xlsx_data[code]['price_manual_may'] = data['price_manual_may']
                                    else:
                                        xlsx_data[code]['price_manual_may'] = float(data['price_manual_may'].strip())
                                    try:
                                        float(sheet[cell].value)
                                        sheet[cell] = xlsx_data[code]['price_manual_may']
                                    except Exception as e:
                                        print(e)
                                        sheet[cell] = '********'
                                        xlsx_data[code]['price_manual_may'] = '*******'
                                    
                                elif is_mi > -1:
                                    if isinstance(data['price_manual_min'], float):
                                        xlsx_data[code]['price_manual_min'] = data['price_manual_min']
                                    else:
                                        xlsx_data[code]['price_manual_min'] = float(data['price_manual_min'].strip())
                                    try:
                                        float(sheet[cell].value)
                                        sheet[cell] = xlsx_data[code]['price_manual_min']
                                    except:
                                        sheet[cell] = '********'
                                        xlsx_data[code]['price_manual_min'] = '*******'
                                else:
                                    print('Error: el archivo tiene q esta contenido en una carpeta con nombre ma o mi')
                            except ValueError:
                                
                                sheet[cell] = '********'
                            
                        else:
                            print(f"{code}: el precio no cambio")
                except Exception as e:
                    print(f"Error: {e}")

                
                to_return.append((code, xlsx_data[code]))

            
        sheet["A1"] = datetime.now().date()
        wb.save(rute_xlsx)
        # copiamos la lista a la carpta listas de precios
        if is_mi > -1:
            folder_name = 'minorista/'
        elif is_ma > -1:
            folder_name = 'mayorista/'
        else:
            folder_name = ''
            print('Error: No se copiaron los archivos. El archivo tiene q esta contenido en una carpeta con nombre ma o mi')
        if folder_name != '':
            path_file = os.path.abspath(f'{XLSX_RUTE}{folder_name}')
            copy_file(rute_xlsx, path_file)

    return [xlsx_name, to_return]

# retorna todos los archivos excel de un directorio
def list_xlsx_to_folder(path):
    archivos_xlsx = []

    for root, dirs, files in os.walk(path):
        for file in files:
            if file.endswith(".xlsx"):
                archivos_xlsx.append(os.path.join(root, file))

    return archivos_xlsx

# recive una lista con todos los registros de listXlsx
# devuelve las listas con los codigos correspondientes
def get_codes_to_xlsx_list(lists_xlsx):
    len_list_xlsx = len(lists_xlsx)
    for i in range(len_list_xlsx,0,-1):
        xlsx = lists_xlsx.pop()
        if xlsx != None:
            artics = ModelArtic.objects.filter(listXlsxID=xlsx).values("code")
            codes = ""
            for artic in artics:
                codes += artic['code'] + ', '

            codes = codes[:-2]
            files_drive = ModelFileDrive.objects.filter(listXlsxID=xlsx)

            links = []
            for file in files_drive:
                if file.parentId.name == "ma":
                    label = "Link comun Mayorista"
                elif file.parentId.name == "mi":
                    label = "Link comun Minorista"
                elif file.parentId.parentId.name == "ma":
                    label = "Link ordenado Mayorista"
                elif file.parentId.parentId.name == "mi":
                    label = "Link ordenado Minorista"
                else:
                    label = "no label"
                links.append((label,f"https://docs.google.com/spreadsheets/d/{file.driveId}/edit#gid=1660344131"))
            lists_xlsx = [(xlsx, codes, links)] + lists_xlsx

    return lists_xlsx


# compara si hay diferencia en los precios y devuelve una lista con todos los archivos xlsx 
# q requieren una actualizacion
def update_artics(artics):
    with transaction.atomic():
        
        db_artics = ModelArtic.objects.all()
        
        to_update = []
        no_changes = []

        
        # verifico q los precios esten actualizados, si no los actualiza
        for code, data in artics.items():
            artic_exist = False
            for artic in db_artics:
                if artic.code == code:
                    try:
                        diff_ma = abs(round(artic.priceMa,1) - round(data['priceMa'],1))
                        diff_mi = abs(round(artic.priceMi,1) - round(data['priceMi'],1))
                        if diff_ma > 0.19  or diff_mi > 0.19:
                            
                            artic.priceMa = data['priceMa']
                            artic.priceMi = data['priceMi']

                        

                            xlsx = ModelToUpdateList.objects.filter(xlsxId=artic.listXlsxID).first()
                            if xlsx == None:
                                xlsx = ModelToUpdateList(xlsxId=artic.listXlsxID)
                                xlsx.save()
                            
                            if not xlsx.xlsxId in to_update:
                                to_update.append(xlsx.xlsxId)
                            
                        elif artic.listXlsxID is not None:
                            
                            if not artic.listXlsxID in no_changes and not artic.listXlsxID in to_update:
                                    # ordeno las listas por fecha
                                    len_no_changes = len(no_changes)
                                    if len_no_changes > 0:
                                        
                                        if no_changes[-1].modDate > artic.listXlsxID.modDate:
                                            no_changes.append(artic.listXlsxID)
                                        else:
                                            for i in range(len_no_changes-1,0,-1):
                                               
                                                if no_changes[i].modDate > artic.listXlsxID.modDate:
                                                    no_changes.insert(i, artic.listXlsxID)
                                                    break
                                            else:
                                                no_changes.insert(0, artic.listXlsxID)
                                    else:
                                        no_changes.append(artic.listXlsxID)

                        if artic.description != data['description']:
                            artic.description = data['description']
                        artic.save()    
                    except KeyError:
                        print(KeyError("Error en la data"))
                    except Exception as e:
                        print("*************************************")
                        print(f"Error xlsxTools.update_artics: {e}")

                    artic_exist = True
                
            # si el articulo no existe lo crea
            if not artic_exist:
                new_artic = ModelArtic(code=code, description=data['description'], priceMa=data['priceMa'], priceMi=data['priceMi'])
                new_artic.save()

            
    to_change_db = ModelToUpdateList.objects.all()
    for xlsx_change in to_change_db:
        if not xlsx_change.xlsxId in to_update:
            to_update.append(xlsx_change.xlsxId)
        try:
            no_changes.remove(xlsx.xlsxId)
        except:
            pass


    to_update = get_codes_to_xlsx_list(to_update)

    no_changes = get_codes_to_xlsx_list(no_changes)

    # esto es un parche porque no entiendo porque se repiten las listas
    # solucion: en vez de utilizar 2 diccionarios hay q usar 1 solo y agregar un campo extra q diga a actualizar 
    for update_xlsx in to_update:
        if update_xlsx in no_changes:
            no_changes.remove(update_xlsx)
    return {
        'to_update': to_update,
        'no_changes': no_changes
        }

if __name__ == '__main__':

    

    results = list_xlsx_to_folder(RUTE_XLSX_AGRUPS['ma'])
    

    print(results)



    
    
